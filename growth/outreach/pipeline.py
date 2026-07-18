"""Buddee outreach pipeline — prospect spreadsheet → ready-to-send batch.

Usage:
    python -m growth.outreach.pipeline \
        --input Buddee_California_Prospect_List2.xlsx \
        --tier A --out growth/outbox

For each prospect row it:
  1. Picks the template by the sheet's "Target Role To Find First" column.
  2. Builds a relevance line from the row's structure/source/fit notes.
  3. Renders first email, LinkedIn note (<=300 chars), and both follow-ups
     with cadence dates (+4 and +8 business days from --send-date).
  4. Runs the claims-discipline linter; refuses to write drafts that fail.
  5. Writes one folder per org plus tracker.csv for the whole batch.

Requires: openpyxl (pip install openpyxl).
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from growth.outreach import templates as T
from growth.outreach.claims_lint import lint

EXPECTED_HEADER = ["Tier", "Organization", "City, State", "Structure / Type", "Source",
                   "Website", "Fit Notes", "Target Role", "Status", "Next Touch", "Notes"]


def slugify(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")[:48]


def add_business_days(start: date, days: int) -> date:
    d = start
    added = 0
    while added < days:
        d += timedelta(days=1)
        if d.weekday() < 5:
            added += 1
    return d


def org_short(org: str) -> str:
    """Compact org name for the 300-char LinkedIn budget."""
    # Drop parenthetical aliases like "(fka Apollo ...)" — history, not the brand.
    org = re.sub(r"\s*\([^)]*\)", "", org).strip()
    cut = re.split(r",| LLC| Inc| Medical Group| Physicians| Management", org)[0].strip()
    return cut if len(cut) >= 4 else org.strip()


def relevance_line(row: dict) -> tuple[str, str]:
    """(full sentence for first email, short clause for follow-up #1)."""

    source = (row.get("source") or "").lower()
    structure = (row.get("structure") or "").lower()
    fit = (row.get("fit_notes") or "").lower()
    if "aco reach" in source or "aco reach" in structure or "reach" in fit:
        return (
            "I'm reaching out because you're on the CMS PY2026 ACO REACH participant "
            "list: under global risk, every documented-but-uncoded HCC lands directly "
            "on your benchmark, so capture accuracy isn't back-office hygiene — it's "
            "the contract's whole economics.",
            "your ACO REACH global-risk contract means every recovered HCC flows "
            "straight to your shared-savings benchmark.",
        )
    if "high-needs" in structure or "home-based" in structure or "chronic" in fit or "high need" in fit:
        return (
            "A home-based, high-needs Medicare panel is the most documentation-intense "
            "chronic-disease mix there is — exactly the encounter type where manual "
            "HCC capture slips most.",
            "your high-needs senior panel is exactly the chronic-disease mix where "
            "manual capture slips most.",
        )
    if "mso" in structure:
        return (
            "As the MSO, you're the one relationship that can put better HCC capture "
            "in front of every IPA you manage at once — without any of them hiring a "
            "coding team.",
            "one MSO relationship puts recovered HCC revenue in front of every IPA "
            "you manage.",
        )
    # Default: independent IPA / medical group (APG-member framing).
    return (
        "Independent groups like yours carry Medicare Advantage risk without a "
        "health-system coding department behind them — which is exactly where "
        "documented-but-uncoded HCCs pile up.",
        "an independent MA-risk panel without a health-system coding department "
        "is where manual capture slips most.",
    )


def read_prospects(path: Path, sheet: str, tier: str | None) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows = []
    header_seen = False
    for raw in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in raw]
        if not header_seen:
            if cells and cells[0].lower() == "tier":
                header_seen = True
            continue
        if not cells or not cells[0]:
            continue
        if tier and cells[0].upper() != tier.upper():
            continue
        rows.append({
            "tier": cells[0], "org": cells[1], "city": cells[2],
            "structure": cells[3], "source": cells[4], "website": cells[5],
            "fit_notes": cells[6], "target_role": cells[7],
        })
    return rows


def render_prospect(row: dict, send_date: date) -> dict:
    role_key = (row["target_role"] or "rcm/ops director").lower()
    if "mso" in row["structure"].lower():
        role_key = "billing/mso"
    tpl_key, email_tpl, subject_tpl = T.ROLE_TEMPLATE_MAP.get(
        role_key, T.ROLE_TEMPLATE_MAP["rcm/ops director"]
    )
    rel_full, rel_short = relevance_line(row)
    fields = {
        "org": row["org"],
        "org_short": org_short(row["org"]),
        "first_name": "[First name]",
        "relevance_line": rel_full,
        "relevance_line_short": rel_short,
        **T.FOUNDER,
    }
    subject = subject_tpl.format(**fields)
    email = email_tpl.format(**fields)
    linkedin = T.LINKEDIN_NOTE.format(**fields)
    fu1 = T.FOLLOWUP_1.format(original_subject=subject, **fields)
    fu2 = T.FOLLOWUP_2.format(original_subject=subject, **fields)
    return {
        "template": tpl_key,
        "subject": subject,
        "email": email,
        "linkedin": linkedin,
        "linkedin_chars": len(linkedin),
        "followup_1": fu1,
        "followup_1_date": add_business_days(send_date, 4),
        "followup_2": fu2,
        "followup_2_date": add_business_days(send_date, 8),
    }


def write_batch(prospects: list[dict], out_dir: Path, send_date: date) -> tuple[int, list[str]]:
    out_dir.mkdir(parents=True, exist_ok=True)
    problems: list[str] = []
    tracker_rows = []
    written = 0
    for i, row in enumerate(prospects, 1):
        r = render_prospect(row, send_date)
        combined = "\n\n".join([r["subject"], r["email"], r["linkedin"], r["followup_1"], r["followup_2"]])
        result = lint(combined, first_touch=True)
        if not result.ok:
            problems.append(f"{row['org']}: BLOCKED — " + "; ".join(result.errors))
            continue
        if r["linkedin_chars"] > 300:
            problems.append(f"{row['org']}: LinkedIn note {r['linkedin_chars']} chars (>300) — shorten org name")
            continue
        folder = out_dir / f"{i:02d}_{slugify(row['org'])}"
        folder.mkdir(parents=True, exist_ok=True)
        (folder / "brief.md").write_text(
            f"# {row['org']}\n\n"
            f"- Location: {row['city']}\n- Structure: {row['structure']}\n"
            f"- Target role: {row['target_role']} (template {r['template']})\n"
            f"- Website: {row['website'] or 'n/a'}\n- Source: {row['source']}\n"
            f"- Fit notes: {row['fit_notes']}\n\n"
            f"Before sending: find the actual {row['target_role']} on LinkedIn, "
            f"replace [First name], verify the fit note still holds, add the demo link.\n",
            encoding="utf-8",
        )
        (folder / "01_first_email.md").write_text(
            f"**Send:** {send_date.isoformat()} (Tue–Thu morning)\n"
            f"**Subject:** {r['subject']}\n\n{r['email']}", encoding="utf-8")
        (folder / "02_linkedin_note.md").write_text(
            f"**Chars:** {r['linkedin_chars']}/300\n\n{r['linkedin']}", encoding="utf-8")
        (folder / "03_followup_1.md").write_text(
            f"**Send:** {r['followup_1_date'].isoformat()} (+4 business days, if no reply)\n"
            f"**Subject:** Re: {r['subject']}\n\n{r['followup_1']}", encoding="utf-8")
        (folder / "04_followup_2.md").write_text(
            f"**Send:** {r['followup_2_date'].isoformat()} (+8 business days, break-up)\n"
            f"**Subject:** Re: {r['subject']}\n\n{r['followup_2']}", encoding="utf-8")
        if result.warnings:
            (folder / "LINT_WARNINGS.md").write_text(
                "\n".join(f"- {w}" for w in result.warnings), encoding="utf-8")
        tracker_rows.append({
            "org": row["org"], "tier": row["tier"], "template": r["template"],
            "target_role": row["target_role"], "city": row["city"],
            "first_send": send_date.isoformat(),
            "followup_1": r["followup_1_date"].isoformat(),
            "followup_2": r["followup_2_date"].isoformat(),
            "status": "drafted", "contact_name": "", "contact_email": "", "reply": "",
        })
        written += 1
    if tracker_rows:
        with open(out_dir / "tracker.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(tracker_rows[0].keys()))
            writer.writeheader()
            writer.writerows(tracker_rows)
    return written, problems


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", required=True, help="Prospect .xlsx (List2 format)")
    p.add_argument("--sheet", default="Prospect List")
    p.add_argument("--tier", default="A")
    p.add_argument("--out", default="growth/outbox")
    p.add_argument("--send-date", default=None, help="YYYY-MM-DD (default: today)")
    p.add_argument("--limit", type=int, default=0)
    args = p.parse_args()

    send_date = date.fromisoformat(args.send_date) if args.send_date else date.today()
    prospects = read_prospects(Path(args.input), args.sheet, args.tier)
    if args.limit:
        prospects = prospects[: args.limit]
    if not prospects:
        sys.exit(f"No tier-{args.tier} prospects found in {args.input}")

    out_dir = Path(args.out) / send_date.isoformat()
    written, problems = write_batch(prospects, out_dir, send_date)
    print(f"Wrote {written}/{len(prospects)} prospect folders to {out_dir}")
    for prob in problems:
        print(f"  !! {prob}")


if __name__ == "__main__":
    main()
